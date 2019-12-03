from openpyxl import load_workbook
import tkinter as tk
from tkinter import filedialog


class OperateExcel(object):
    def __init__(self):
        self.wb = None
        self.sheet = None
        self.info_dict = {}

    @staticmethod
    def read_file():
        root = tk.Tk()
        root.withdraw()
        file_names = filedialog.askopenfilenames(filetypes=[('xlsx', '*.xlsx'), ])
        return file_names[0]

    def data_cleaning(self, filename):
        self.wb = load_workbook(filename)
        self.sheet = self.wb["销售订单汇总表"]
        names = self.sheet['D']
        custom_record_list = [name.value for name in names]
        custom_name_list = list(set(custom_record_list))
        custom_name_list.remove(None)
        custom_name_list.remove('客户名称')
        for name in custom_name_list:
            if custom_record_list.count(name) > 1:
                self.info_dict[name] = [cus_index for cus_index, cus_value in enumerate(custom_record_list) if
                                        cus_value == name]

    def save_excel(self):
        for key, value in self.info_dict.items():
            data = []
            for index in value:
                items = []
                for r in range(self.sheet.max_row):
                    if index == r:
                        for c in range(self.sheet.max_column):
                            items.append(self.sheet.cell(r + 1, c + 1).value)
                data.append(items)
            new_sheet = self.wb.create_sheet(key)
            prefix = ['序号', '日期', '订单号', '客户名称', '客户分类', '客户编码', '结账方式', '车型', '车架号', '施工内容', '材料代码', '数量', '单价', '总金额',
                      '优惠金额',
                      '实际销售额', '已收定金', '已收尾款', '手续费', '应收账款', '是否清款', '发货日期']
            new_sheet.append(prefix)
            for value_list in data:
                new_sheet.append(value_list)
        self.wb.save('销售订单汇总表1.xlsx')

    def run(self):
        filename = self.read_file()
        self.data_cleaning(filename)
        self.save_excel()


if __name__ == '__main__':
    excel_obj = OperateExcel()
    excel_obj.run()
